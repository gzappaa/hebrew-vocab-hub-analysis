"""
Hebrew Vocab Hub — Excel/Power BI Data Export
================================================
Generates one workbook (hebrew_vocab_hub_report.xlsx) with 8 sheets,
each mapped 1:1 to an analysis already done in the notebooks. This
workbook is meant to be BOTH your Excel deliverable and the data
source you point Power BI Desktop at.


Requirements: pandas, sqlalchemy, psycopg2, python-dotenv, openpyxl
    pip install pandas sqlalchemy psycopg2-binary python-dotenv openpyxl

Expects the same .env vars as your notebooks:
    DB_USER, DB_PASSWORD, DB_HOST, DB_PORT, DB_NAME

Expects (same as 02_source_comparison.ipynb):
    ../data/stopwords.txt
    ../data/word_semantic_labels.csv   (already generated — not re-called here)
"""

import os
import numpy as np
import pandas as pd
from sqlalchemy import create_engine, text
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font

load_dotenv()

engine = create_engine(
    f"postgresql+psycopg2://{os.getenv('DB_USER')}:{os.getenv('DB_PASSWORD')}"
    f"@{os.getenv('DB_HOST')}:{os.getenv('DB_PORT')}/{os.getenv('DB_NAME')}"
)

STOPWORDS_PATH = "../data/stopwords.txt"
LABELS_PATH = "../data/word_semantic_labels.csv"
OUTPUT_PATH = "../excel/hebrew_vocab_hub_report.xlsx"


# ------------------------------------------------------------
# Sheet 1 — Overview (baseline row counts)
# ------------------------------------------------------------
def build_overview():
    tables = ["roots", "lemmas", "words", "word_lemmas",
              "conj_tables", "conj_cells", "sentences", "word_sources"]
    rows = []
    for t in tables:
        count = pd.read_sql(f"SELECT COUNT(*) FROM {t}", engine).iloc[0, 0]
        rows.append({"table_name": t, "row_count": count})
    return pd.DataFrame(rows)


# ------------------------------------------------------------
# Sheet 2 — POS Distribution (same grouping logic as 01_eda)
# ------------------------------------------------------------
def build_pos_distribution():
    pos_df = pd.read_sql(
        """
        SELECT part_of_speech_plain, COUNT(*) AS n
        FROM lemmas
        WHERE part_of_speech_plain IS NOT NULL
        GROUP BY part_of_speech_plain
        ORDER BY n DESC
        """,
        engine,
    )
    pos_df["part_of_speech_plain"] = pos_df["part_of_speech_plain"].str.strip()

    for prefix in ["noun", "verb", "adjective", "adverb", "preposition",
                   "pronoun", "cardinal numeral"]:
        mask = pos_df["part_of_speech_plain"].str.startswith(prefix, na=False)
        pos_df.loc[mask, "part_of_speech_plain"] = prefix

    pos_df = (
        pos_df.groupby("part_of_speech_plain", as_index=False)["n"]
        .sum()
        .sort_values(by="n", ascending=False)
    )
    pos_df = pos_df[pos_df["part_of_speech_plain"].str.strip() != ""]

    total_n = pos_df["n"].sum()
    threshold = 0.02 * total_n
    major_df = pos_df[pos_df["n"] >= threshold].copy()
    minor_df = pos_df[pos_df["n"] < threshold]

    if not minor_df.empty:
        other_row = pd.DataFrame([{"part_of_speech_plain": "Other", "n": minor_df["n"].sum()}])
        plot_df = pd.concat([major_df, other_row], ignore_index=True)
    else:
        plot_df = major_df

    plot_df = plot_df.rename(columns={"part_of_speech_plain": "part_of_speech", "n": "count"})
    plot_df["pct_of_total"] = (plot_df["count"] / total_n * 100).round(1)
    return plot_df.reset_index(drop=True)


# ------------------------------------------------------------
# Sheet 3 — Top 30 Roots
# ------------------------------------------------------------
def build_top_roots():
    return pd.read_sql(
        """
        SELECT r.display AS root, COUNT(l.id) AS n_lemmas
        FROM roots r
        JOIN lemmas l ON l.root_id = r.id
        GROUP BY r.display
        ORDER BY n_lemmas DESC
        LIMIT 30
        """,
        engine,
    )


# ------------------------------------------------------------
# Sheet 4 — Word Frequency Top 20
# ------------------------------------------------------------
def build_word_frequency_top20():
    df = pd.read_sql(
        """
        SELECT w.word, ws.total
        FROM words w
        JOIN word_sources ws ON ws.word_id = w.id
        WHERE ws.total > 0
        ORDER BY ws.total DESC
        LIMIT 20
        """,
        engine,
    )
    df.insert(0, "rank", range(1, len(df) + 1))
    return df



# ------------------------------------------------------------
# Sheet 5 — Top 20 Words Overall (stopword-filtered, not split by source)
# ------------------------------------------------------------
def build_top_words_overall_filtered():
    query = """
    SELECT w.word, ws.total
    FROM word_sources ws
    JOIN words w ON ws.word_id = w.id
    WHERE ws.total > 0
    """
    df = pd.read_sql(query, engine)

    with open(STOPWORDS_PATH, "r", encoding="utf-8") as f:
        stopwords = set(line.strip() for line in f if line.strip())

    df_filtered = df[~df["word"].isin(stopwords)]
    top = df_filtered.sort_values("total", ascending=False).head(20).reset_index(drop=True)
    top.insert(0, "rank", range(1, len(top) + 1))
    return top


# ------------------------------------------------------------
# Sheet 6 — Source Overlap (venn segment counts)
# ------------------------------------------------------------
def build_source_overlap():
    df = pd.read_sql(
        """
        SELECT ws.word_id, ws.songs, ws.news, ws.youtube
        FROM word_sources ws
        """,
        engine,
    )
    in_songs = df["songs"] > 0
    in_news = df["news"] > 0
    in_youtube = df["youtube"] > 0

    segments = {
        "Songs only": (in_songs & ~in_news & ~in_youtube).sum(),
        "News only": (~in_songs & in_news & ~in_youtube).sum(),
        "YouTube only": (~in_songs & ~in_news & in_youtube).sum(),
        "Songs & News only": (in_songs & in_news & ~in_youtube).sum(),
        "Songs & YouTube only": (in_songs & ~in_news & in_youtube).sum(),
        "News & YouTube only": (~in_songs & in_news & in_youtube).sum(),
        "All three sources": (in_songs & in_news & in_youtube).sum(),
    }
    return pd.DataFrame(
        [{"segment": k, "word_count": int(v)} for k, v in segments.items()]
    )


# ------------------------------------------------------------
# Sheet 7 — Top Words By Source (stopword-filtered, top 20 each)
# ------------------------------------------------------------
def build_top_words_by_source():
    query = """
    SELECT ws.word_id, w.word, ws.songs, ws.news, ws.youtube, ws.total
    FROM word_sources ws
    JOIN words w ON ws.word_id = w.id
    """
    df_sources = pd.read_sql(query, engine)

    with open(STOPWORDS_PATH, "r", encoding="utf-8") as f:
        stopwords = set(line.strip() for line in f if line.strip())

    df_filtered = df_sources[~df_sources["word"].isin(stopwords)]

    frames = []
    for col in ["songs", "news", "youtube"]:
        top = (
            df_filtered.sort_values(col, ascending=False)[["word", col]]
            .head(20)
            .rename(columns={col: "freq"})
        )
        top["source"] = col
        frames.append(top)

    return pd.concat(frames, ignore_index=True)[["source", "word", "freq"]]


# ------------------------------------------------------------
# Sheet 8 — Semantic Labels By Source (reuses cached label CSV)
# ------------------------------------------------------------
def build_semantic_labels_by_source():
    if not os.path.exists(LABELS_PATH):
        print(f"Warning: {LABELS_PATH} not found — skipping Semantic_Labels_By_Source sheet.")
        return pd.DataFrame(columns=["source", "label", "count"])

    query = """
    SELECT ws.word_id, w.word, ws.songs, ws.news, ws.youtube
    FROM word_sources ws
    JOIN words w ON ws.word_id = w.id
    """
    df_sources = pd.read_sql(query, engine)

    with open(STOPWORDS_PATH, "r", encoding="utf-8") as f:
        stopwords = set(line.strip() for line in f if line.strip())
    df_filtered = df_sources[~df_sources["word"].isin(stopwords)]

    frames = []
    for col in ["songs", "news", "youtube"]:
        top = (
            df_filtered.sort_values(col, ascending=False)[["word", col]]
            .head(50)
            .rename(columns={col: "freq"})
        )
        top["source"] = col
        frames.append(top)
    top_words_combined = pd.concat(frames, ignore_index=True).reset_index(drop=True)
    top_words_combined["row_id"] = top_words_combined.index

    df_labels = pd.read_csv(LABELS_PATH)
    top_words_combined = top_words_combined.merge(
        df_labels[["row_id", "label"]], on="row_id", how="left"
    )

    return (
        top_words_combined.groupby(["source", "label"])
        .size()
        .reset_index(name="count")
    )


# ------------------------------------------------------------
# Sheet 9 — Frequency By Source Coverage
# ------------------------------------------------------------
def build_frequency_by_coverage():
    df = pd.read_sql(
        """
        SELECT w.word, ws.songs, ws.news, ws.youtube, ws.total
        FROM word_sources ws
        JOIN words w ON ws.word_id = w.id
        WHERE ws.total > 0
        """,
        engine,
    )
    df["source_count"] = (
        (df["songs"] > 0).astype(int)
        + (df["news"] > 0).astype(int)
        + (df["youtube"] > 0).astype(int)
    )
    df["log_total"] = np.log1p(df["total"])
    return df[["word", "total", "source_count", "log_total"]]


# ------------------------------------------------------------
# Write workbook
# ------------------------------------------------------------
def main():
    sheets = {
        "Overview": build_overview(),
        "POS_Distribution": build_pos_distribution(),
        "Top_Roots": build_top_roots(),
        "Word_Frequency_Top20": build_word_frequency_top20(),
        "Filtered_Word_Frequency_Top20": build_top_words_overall_filtered(),
        "Source_Overlap": build_source_overlap(),
        "Top_Words_By_Source": build_top_words_by_source(),
        "Semantic_Labels_By_Source": build_semantic_labels_by_source(),
        "Frequency_By_Coverage": build_frequency_by_coverage(),
    }

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)

    # Light formatting pass: bold headers, professional font, autofit-ish widths
    wb = load_workbook(OUTPUT_PATH)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(name="Arial", bold=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial")
        for col_cells in ws.columns:
            length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            col_letter = col_cells[0].column_letter
            ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 40)
    wb.save(OUTPUT_PATH)

    print(f"Done. Wrote {len(sheets)} sheets to {OUTPUT_PATH}")
    for name, df in sheets.items():
        print(f"  {name}: {len(df)} rows")


if __name__ == "__main__":
    main()